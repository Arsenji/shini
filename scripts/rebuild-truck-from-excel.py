#!/usr/bin/env python3
"""Синхронизация грузовых шин из «Прайс лист для сайта-2.xlsx» → products.ts."""

from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PRODUCTS_TS = ROOT / "src/data/shop/products.ts"
DEFAULT_XLSX = Path.home() / "Downloads/Прайс лист для сайта-2.xlsx"

SPEC_HINT = re.compile(
    r"^(?:\d+/\d+[A-Za-zА-Яа-я]?|"
    r"\d+(?:сл|сл\.|PR|pr|нс\.|НС\.|TT|TL|"
    r"а/шина|универ\.|Б[Мм]|ОМСК|Омск|Н/К|н/к).*)",
    re.I,
)


def normalize_size(raw: str) -> str:
    s = re.sub(r"\s+", "", raw.replace(",", "."))
    m = re.match(r"^(\d+(?:\.\d+)?)[Rr](\d+(?:\.\d+)?)$", s)
    if m:
        return f"{m.group(1)}R{m.group(2)}"
    return s.upper()


def slugify(text: str) -> str:
    import hashlib

    t = text.lower().replace("\\", "/")
    t = re.sub(r"[х×]", "x", t)
    t = re.sub(r"\s+", "-", t.strip())
    ascii_part = re.sub(r"[^a-z0-9._-]+", "", t, flags=re.I)
    if ascii_part:
        return ascii_part[:48].strip("-")
    return hashlib.sha1(text.encode("utf-8")).hexdigest()[:10]


def looks_like_model_token(token: str) -> bool:
    if SPEC_HINT.match(token):
        return False
    if re.match(r"^нс\.?\d+", token, re.I):
        return False
    if re.match(r"^\d+сл", token, re.I):
        return False
    if re.match(r"^[A-Za-zА-Яа-я]{2,}[- ]?[A-Za-z0-9./]+$", token):
        return True
    if re.match(r"^[A-Z]{2,}\d+", token):
        return True
    return False


def parse_truck_line(name: str) -> tuple[str, str, str, str]:
    """size, brand, model, truck_specs (слойность, индекс нагрузки и пр.)"""
    name = " ".join(name.split()).replace("\\", "/")

    m = re.match(r"^([\d.]+\s*R[\d.]+)\s+(.+)$", name, re.I)
    if m:
        size = normalize_size(m.group(1))
        rest = m.group(2).strip()
        tokens = rest.split()
        if len(tokens) >= 2 and looks_like_model_token(tokens[1]):
            brand = tokens[0]
            model = tokens[1]
            specs = " ".join(tokens[2:])
        else:
            brand = tokens[0] if tokens else rest
            model = size
            specs = " ".join(tokens[1:]) if len(tokens) > 1 else ""
        return size, brand, model, specs.strip()

    m = re.match(r"^(.+?)\s+([\d.]+\s*R[\d.]+)\s*(.*)$", name, re.I)
    if m:
        brand = m.group(1).strip()
        size = normalize_size(m.group(2))
        specs = m.group(3).strip()
        return size, brand, size, specs

    return name, name, name, ""


def load_excel_truck_rows(xlsx: Path) -> list[tuple[str, int, str, str, str, str]]:
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    ws = openpyxl.load_workbook(xlsx, data_only=True)["Лист1"]
    truck = False
    rows: list[tuple[str, int, str, str, str, str]] = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, 1).value
        if cell == "грузовые":
            truck = True
            continue
        if truck and isinstance(cell, str) and cell.startswith("диски"):
            break
        if not truck:
            continue
        price = ws.cell(r, 2).value
        if not isinstance(price, (int, float)):
            continue
        raw = str(cell).strip()
        size, brand, model, specs = parse_truck_line(raw)
        rows.append((raw, int(price), size, brand, model, specs))
    return rows


def load_non_truck_products() -> list[dict]:
    cmd = [
        "npx",
        "tsx",
        "-e",
        """
import { shopProducts } from './src/data/shop/products.ts'
console.log(JSON.stringify(shopProducts.filter(p => p.category !== 'truck')))
""",
    ]
    out = subprocess.check_output(cmd, cwd=ROOT, text=True)
    return json.loads(out)


def load_existing_truck_ids() -> dict[tuple[str, str, str, int], str]:
    cmd = [
        "npx",
        "tsx",
        "-e",
        """
import { shopProducts } from './src/data/shop/products.ts'
const truck = shopProducts.filter(p => p.category === 'truck')
console.log(JSON.stringify(truck.map(p => ({
  id: p.id,
  brand: p.brand,
  model: p.model,
  size: p.sizeGroup || p.sizes[0] || '',
  price: p.price ?? 0,
}))))
""",
    ]
    out = subprocess.check_output(cmd, cwd=ROOT, text=True)
    items = json.loads(out)
    mapping: dict[tuple[str, str, str, int], str] = {}
    for item in items:
        key = (item["brand"], item["model"], item["size"], int(item["price"]))
        if key not in mapping or (item["id"].endswith("-2") is False and mapping[key].endswith("-2")):
            mapping[key] = item["id"]
    return mapping


def make_id(
    brand: str,
    model: str,
    size: str,
    specs: str,
    price: int,
    legacy: dict[tuple[str, str, str, int], str],
    used: set[str],
) -> str:
    legacy_key = (brand, model, size, price)
    if legacy_key in legacy and legacy[legacy_key] not in used:
        candidate = legacy[legacy_key]
        used.add(candidate)
        return candidate

    parts = [slugify(brand), slugify(model if model != size else ""), slugify(specs), slugify(size)]
    base = "-".join(p for p in parts if p)[:55].strip("-") or slugify(size)
    candidate = base
    n = 2
    while candidate in used:
        candidate = f"{base}-{n}"
        n += 1
    used.add(candidate)
    return candidate


def ts_string(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"


def render_product(p: dict) -> str:
    body = [
        f"    id: {ts_string(p['id'])},",
        f"    brand: {ts_string(p['brand'])},",
        f"    model: {ts_string(p['model'])},",
        "    category: 'truck',",
        "    imageKey: 'truck',",
        f"    sizeGroup: {ts_string(p['sizeGroup'])},",
        f"    sizes: [{ts_string(p['sizes'][0])}],",
    ]
    if p.get("truckSpecs"):
        body.append(f"    truckSpecs: {ts_string(p['truckSpecs'])},,")
    body.extend(
        [
            f"    price: {p['price']},",
            "    offers: [",
            f"      {{ size: {ts_string(p['sizes'][0])}, price: {p['price']} }},",
            "    ],",
        ]
    )
    body = [line.replace(",,", ",") for line in body]
    return "  {\n" + "\n".join(body) + "\n  }"


def write_products_ts(non_truck: list[dict], truck: list[dict]) -> None:
    header = """import type { ShopProduct } from './types'

/**
 * Каталог магазина (автогенерация грузовых: scripts/rebuild-truck-from-excel.py).
 */
export const shopProducts: ShopProduct[] = [
"""
    # Re-export non-truck via tsx formatter
    dump = json.dumps(non_truck, ensure_ascii=False)
    node = f"""
const nonTruck = {dump}
function tsStr(s) {{
  if (s == null) return 'null'
  return "'" + String(s).replace(/\\\\/g, '\\\\\\\\').replace(/'/g, "\\\\'") + "'"
}}
function renderOffer(o) {{
  return '{{ size: ' + tsStr(o.size) + ', price: ' + o.price + ' }}'
}}
function renderProduct(p) {{
  const lines = ['  {{', '    id: ' + tsStr(p.id) + ',', '    brand: ' + tsStr(p.brand) + ',', '    model: ' + tsStr(p.model) + ',']
  if (p.category) lines.push('    category: ' + tsStr(p.category) + ',')
  if (p.season) lines.push('    season: ' + tsStr(p.season) + ',')
  if (p.imageKey) lines.push('    imageKey: ' + tsStr(p.imageKey) + ',')
  if (p.sizeGroup) lines.push('    sizeGroup: ' + tsStr(p.sizeGroup) + ',')
  if (p.sizes?.length) lines.push('    sizes: [' + p.sizes.map(tsStr).join(', ') + '],')
  if (p.badge) lines.push('    badge: ' + tsStr(p.badge) + ',')
  if (p.image) lines.push('    image: ' + tsStr(p.image) + ',')
  if (p.color) lines.push('    color: ' + tsStr(p.color) + ',')
  if (p.truckSpecs) lines.push('    truckSpecs: ' + tsStr(p.truckSpecs) + ',')
  if (typeof p.price === 'number') lines.push('    price: ' + p.price + ',')
  if (p.offers?.length) {{
    lines.push('    offers: [')
    for (const o of p.offers) lines.push('      ' + renderOffer(o) + ',')
    lines.push('    ],')
  }}
  lines.push('  }}')
  return lines.join('\\n')
}}
console.log(nonTruck.map(renderProduct).join(',\\n'))
"""
    out = subprocess.check_output(["npx", "tsx", "-e", node], cwd=ROOT, text=True)
    truck_blocks = ",\n".join(render_product(p) for p in truck)
    content = header + out.rstrip() + ",\n" + truck_blocks + "\n]\n"
    PRODUCTS_TS.write_text(content, encoding="utf-8")


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Excel not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    rows = load_excel_truck_rows(xlsx)
    legacy = load_existing_truck_ids()
    used_ids: set[str] = set()
    truck_products: list[dict] = []

    for raw, price, size, brand, model, specs in rows:
        pid = make_id(brand, model, size, specs, price, legacy, used_ids)
        item: dict = {
            "id": pid,
            "brand": brand,
            "model": model,
            "sizeGroup": size,
            "sizes": [size],
            "price": price,
        }
        if specs:
            item["truckSpecs"] = specs
        truck_products.append(item)

    non_truck = load_non_truck_products()
    write_products_ts(non_truck, truck_products)
    print(f"Wrote {len(non_truck)} non-truck + {len(truck_products)} truck → {PRODUCTS_TS}")


if __name__ == "__main__":
    main()
